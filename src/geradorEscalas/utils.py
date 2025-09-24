# utilitarios.py

from datetime import datetime
import calendar
from openpyxl.utils import get_column_letter
import re

def preencher_escala(ws, linha_inicio, colaboradores, dias_mes, dias_semana, setor_nome, novos_turnos, afastados, mes, ano):
    colaboradores.sort(key=lambda x: str(x.get("Nome", "")).strip().lower())
    for i, colab in enumerate(colaboradores):
        linha = linha_inicio + 1 + i
        nome = colab.get("Nome", "")
        matricula = str(colab.get("Matrícula", "")).strip()
        escala = str(colab.get("Escala", "")).strip().lower()
        turno_k = str(colab.get('Resultado Esperado', '') or '').strip().lower()
        turno_f = str(colab.get('Tipo de Turno', '')).strip().lower()
        turno = turno_k if turno_k and turno_k != 'nan' else turno_f
        ws[f"A{linha}"] = nome
        ws[f"B{linha}"] = colab.get("Cargo", "")
        ws[f"C{linha}"] = matricula
        ws[f"D{linha}"] = colab.get("COREN (opcional)", "")
        ws[f"E{linha}"] = colab.get("Horário Padrão", "")
        estado_dia = True if turno in ['diurno 1', 'noturno 1'] else False
        for dia in range(1, dias_mes + 1):
            col = get_column_letter(6 + dia - 1)
            celula = f"{col}{linha}"
            dia_semana = dias_semana[dia - 1]
            if escala == "12x36":
                ws[celula] = "X" if estado_dia else "F"
                estado_dia = not estado_dia
            elif escala == "diarista":
                ws[celula] = "X" if dia_semana in ["SEG", "TER", "QUA", "QUI", "SEX"] else "F"
        afastamento_info = afastados.get(matricula)
        if afastamento_info:
            inicio, fim, motivo = afastamento_info
            for dia in range(1, dias_mes + 1):
                data_atual = datetime(ano, mes, dia)
                if inicio <= data_atual <= fim:
                    col = get_column_letter(6 + dia - 1)
                    ws[f"{col}{linha}"] = motivo
        novos_turnos[matricula] = {"nome": nome, "setor": setor_nome, "turno": turno, "matricula": matricula}

def atualizar_turnos_com_logica_prioritaria(wb, mes, ano):
    aba = wb["Cadastro_Colaboradores"]
    headers = {cell.value: idx for idx, cell in enumerate(aba[1], 1)}
    ultimo_dia_mes = calendar.monthrange(ano, mes)[1]
    dia_impar = ultimo_dia_mes % 2 != 0
    for linha in range(2, aba.max_row + 1):
        resultado_k = aba.cell(row=linha, column=headers["Resultado Esperado"]).value
        tipo_turno_atual = aba.cell(row=linha, column=headers["Tipo de Turno"]).value
        resultado_k_str = str(resultado_k).strip().lower() if resultado_k else ""
        tipo_turno_str = str(tipo_turno_atual).strip().lower() if tipo_turno_atual else ""
        novo_turno = resultado_k_str if resultado_k_str and resultado_k_str != 'nan' else tipo_turno_str
        aba.cell(row=linha, column=headers["Tipo de Turno"]).value = novo_turno.title()
        novo_k = ""
        if "diurno" in novo_turno or "noturno" in novo_turno:
            base = "diurno" if "diurno" in novo_turno else "noturno"
            if dia_impar: novo_k = f"{base} 2" if "1" in novo_turno else f"{base} 1"
            else: novo_k = novo_turno
        aba.cell(row=linha, column=headers["Resultado Esperado"]).value = novo_k.title() if novo_k else ""

def encontrar_titulos(ws):
    titulos = []
    for row in range(1, ws.max_row + 1):
        valor = str(ws[f"A{row}"].value or "").strip().lower()
        if valor.startswith("título"):
            titulos.append((valor, row))
    return titulos

def ocultar_blocos_vazios(ws, titulos):
    for i in range(len(titulos)):
        linha_titulo = titulos[i][1]
        proxima_linha = titulos[i + 1][1] if i + 1 < len(titulos) else ws.max_row + 1
        bloco_tem_conteudo = any(ws[f"A{linha}"].value for linha in range(linha_titulo + 1, proxima_linha))
        if not bloco_tem_conteudo:
            for linha in range(linha_titulo, proxima_linha):
                ws.row_dimensions[linha].hidden = True

def limpar_cadastro_colaboradores(wb, setores):
    ws = wb["Cadastro_Colaboradores"]
    headers = [cell.value for cell in ws[1]]
    setores_lower = [s.lower().strip() for s in setores]
    linhas_a_manter = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[headers.index("Setor")]).strip().lower() in setores_lower:
            linhas_a_manter.append(row)
    ws.delete_rows(2, ws.max_row)
    for i, linha_dados in enumerate(linhas_a_manter, start=2):
        for j, valor in enumerate(linha_dados, start=1):
            ws.cell(row=i, column=j, value=valor)

def limpar_nome_para_arquivo(nome):
    return re.sub(r'[\\/:"*?<>|]+', '_', str(nome)).strip()