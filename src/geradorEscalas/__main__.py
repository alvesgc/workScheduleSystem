# main.py

import tkinter as tk
from datetime import datetime
import calendar
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# Importa a classe da interface e as funções de utilitários
from interface_moderna import WizardApp
import utilitarios as util

def run_business_logic(data):
    """
    Função com a lógica de negócio, chamada ao final do assistente.
    Retorna (True, mensagem_sucesso) ou (False, mensagem_erro).
    """
    try:
        caminho_modelo = data['caminho_modelo']
        pasta_destino = data['pasta_destino']
        modelo_tipo = data['modelo_tipo']
        mes = data['mes']
        ano = data['ano']
        setores_selecionados = data['setores_selecionados']
        afastados_com_motivo = data.get('afastados_com_motivo', {})
        df_colab_ativos = data['df_colab_ativos']

        afastados_final = {}
        if data.get('afastados_preview'):
            for matricula, nome in data['afastados_preview'].items():
                motivo = afastados_com_motivo.get(matricula, "AFAST")
                row = df_colab_ativos[df_colab_ativos["Matrícula"].astype(str) == matricula].iloc[0]
                periodo_str = str(row.get("Período de Afastamento", "")).strip()
                inicio_str, fim_str = periodo_str.split(" a ")
                inicio = datetime.strptime(inicio_str.strip(), "%d/%m/%Y")
                fim = datetime.strptime(fim_str.strip(), "%d/%m/%Y")
                afastados_final[matricula] = (inicio, fim, motivo)

        nome_mes = calendar.month_name[mes].upper()
        dias_mes = calendar.monthrange(ano, mes)[1]
        dias_semana_pt = ["SEG", "TER", "QUA", "QUI", "SEX", "SAB", "DOM"]
        dias_semana = [dias_semana_pt[calendar.weekday(ano, mes, d)] for d in range(1, dias_mes + 1)]

        wb = load_workbook(caminho_modelo)
        util.atualizar_turnos_com_logica_prioritaria(wb, mes, ano)
        df_colab = df_colab_ativos[df_colab_ativos['Setor'].isin(setores_selecionados)].copy()
        
        aba_base = "ESCALA MODELO MISTO" if modelo_tipo == "1" else "ESCALA MODELO DIARISTA"
        ws_modelo = wb[aba_base]
        ws_escala = wb.copy_worksheet(ws_modelo)
        ws_escala.title = f"ESCALA_{nome_mes}"
        
        novos_turnos = {}
        
        def obter_turno_real(row):
            escala = str(row.get('Escala', '') or '').strip().lower()
            if escala == "diarista": return "diarista"
            turno_k = str(row.get('Resultado Esperado', '') or '').strip().lower()
            turno_f = str(row.get('Tipo de Turno', '')).strip().lower()
            return turno_k if turno_k and turno_k != 'nan' else turno_f

        mapa = {"título 1": "diarista", "título diurno 1": "diurno 1", "título diurno 2": "diurno 2", "título noturno 1": "noturno 1", "título noturno 2": "noturno 2"}
        titulos = util.encontrar_titulos(ws_escala)
        setor_turno_colabs = {}
        for _, row in df_colab.iterrows():
            setor = str(row["Setor"]).strip().lower()
            turno = obter_turno_real(row)
            if (setor, turno) not in setor_turno_colabs: setor_turno_colabs[(setor, turno)] = []
            setor_turno_colabs[(setor, turno)].append(row.to_dict())

        # Lógica de preenchimento revisada para evitar sobreposição
        for titulo_str, linha_base in titulos:
            grupo = mapa.get(titulo_str)
            if not grupo: continue
            
            linha_atual = linha_base + 1
            conteudo_bloco = False
            for setor in setores_selecionados:
                colabs = setor_turno_colabs.get((setor.lower(), grupo), [])
                if colabs:
                    conteudo_bloco = True
                    ws_escala[f'A{linha_atual}'] = setor.upper()
                    util.preencher_escala(ws_escala, linha_atual, colabs, dias_mes, dias_semana, setor, novos_turnos, afastados_final, mes, ano)
                    linha_atual += len(colabs) + 1 # Pula para o próximo espaço
            
            # Limpa o título genérico e mantém apenas se tiver conteúdo
            if conteudo_bloco:
                ws_escala[f'A{linha_base}'].value = f"GRUPO - {grupo.upper()}"
            else:
                ws_escala[f'A{linha_base}'].value = None

        util.ocultar_blocos_vazios(ws_escala, titulos)
        util.limpar_cadastro_colaboradores(wb, setores_selecionados)

        nome_final = f"ESCALA_{nome_mes}_{ano}_{util.limpar_nome_para_arquivo('_'.join(setores_selecionados))}.xlsx"
        caminho_final = os.path.join(pasta_destino, nome_final)
        wb.save(caminho_final)
        
        return True, f"Escala gerada com sucesso!\nSalva em: {caminho_final}"

    except Exception as e:
        # Levanta o erro para ser capturado pela UI
        raise e

if __name__ == "__main__":
    root = tk.Tk()
    app = WizardApp(root, business_logic_callback=run_business_logic)
    root.mainloop()